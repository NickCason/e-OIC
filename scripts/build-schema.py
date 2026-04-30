#!/usr/bin/env python3
"""
build-schema.py — Parse the e-OIC (eTechGroup Onsite Investigation Checklist)
Excel template into a JSON schema the React app consumes.

Usage:
    python3 scripts/build-schema.py public/template.xlsx > src/schema.json

Run this once whenever you replace public/template.xlsx with a new revision.
Then bump VERSION in public/service-worker.js so installed clients pick up
the new template + schema.

Schema shape (per sheet):
    {
      sheet_name, header_row, first_data_row,
      columns: [{ index, group, header }, ...],
      photo_checklist_columns: [str, ...],      // panel-level shot list
      hyperlink_column: str | null,             // auto-filled at export
      primary_key: str,                         // used for row labels & photo folder names
      row_photos_enabled: bool
    }
"""
import sys
import json
import openpyxl

DATA_SHEETS = [
    'Panels', 'Power', 'PLC Racks', 'PLC Slots', 'Fieldbus IO',
    'Network Devices', 'HMIs', 'Ethernet Switches', 'Drive Parameters',
    'Conv. Speeds', 'Safety Circuit', 'Safety Devices', 'Peer to Peer Comms',
]

# Hand-picked primary key column per sheet — what shows in the row picker AND
# becomes the per-row photo folder name. Must match a column header exactly.
PRIMARY_KEY = {
    'Panels':              'Panel Name',
    'Power':               'Device Name',
    'PLC Racks':           'Rack Name',
    'PLC Slots':           'Slot',
    'Fieldbus IO':         'Device Name',
    'Network Devices':     'Device Name',
    'HMIs':                'HMI Name',
    'Ethernet Switches':   'Name',
    'Drive Parameters':    'Device Name',
    'Conv. Speeds':        'Device Name',
    'Safety Circuit':      'Circuit Name',
    'Safety Devices':      'Device Name',
    'Peer to Peer Comms':  'Device Name',
}

# Sheets that allow row-level photos (in addition to panel-level Photo Checklist).
ROW_PHOTOS_ENABLED = set(DATA_SHEETS)


def build(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    schema = {}

    for name in DATA_SHEETS:
        if name not in wb.sheetnames:
            print(f'!! Missing sheet: {name}', file=sys.stderr)
            continue
        ws = wb[name]
        max_col = ws.max_column
        r1 = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        r2 = [ws.cell(row=2, column=c).value for c in range(1, max_col + 1)]
        r3 = [ws.cell(row=3, column=c).value for c in range(1, max_col + 1)]

        # Network Devices: row 1 is description, row 2 groups, row 3 headers, row 4 example.
        # All other sheets: row 1 groups, row 2 headers, row 3 example.
        if name == 'Network Devices':
            groups_row, headers_row, header_excel_row, first_data_row = r2, r3, 3, 4
        else:
            groups_row, headers_row, header_excel_row, first_data_row = r1, r2, 2, 3

        filled_groups, last = [], None
        for g in groups_row:
            if g is not None:
                last = g
            filled_groups.append(last)

        columns = []
        for i, header in enumerate(headers_row):
            if header is None:
                continue
            columns.append({
                'index': i,
                'group': filled_groups[i],
                'header': str(header).replace('\n', ' ').strip(),
            })

        photo_cols = [c['header'] for c in columns if c['group'] == 'Photo Checklist']
        link_col = next((c['header'] for c in columns if 'Hyperlink' in c['header']), None)

        pk = PRIMARY_KEY.get(name)
        if not any(c['header'] == pk for c in columns):
            print(f'!! PK "{pk}" not found in {name}, falling back', file=sys.stderr)
            pk = next(
                (c['header'] for c in columns
                 if 'Hyperlink' not in c['header']
                 and not c['header'].endswith(('Completed', 'Complete', 'Uploaded', 'Backup'))),
                None
            )

        schema[name] = {
            'sheet_name': name,
            'header_row': header_excel_row,
            'first_data_row': first_data_row,
            'columns': columns,
            'photo_checklist_columns': photo_cols,
            'hyperlink_column': link_col,
            'primary_key': pk,
            'row_photos_enabled': name in ROW_PHOTOS_ENABLED,
        }

    return schema


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: build-schema.py <template.xlsx>', file=sys.stderr)
        sys.exit(1)
    print(json.dumps(build(sys.argv[1]), indent=2))
